import json
import os
from pathlib import Path
import numpy as np
import pandas as pd
from ..BVP.ABS_Column import abs_column
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt


# Put outputs into dictionary and dataframe

def make_dfs_dict(output_dict, keys_dict, stages, coordinate_frame=None):
    sheetnames = list(keys_dict.keys())
    dfs_dict = {}
    coordinate_frame = coordinate_frame.reset_index(drop=True) if coordinate_frame is not None else None
    for k1 in sheetnames:
        d = {}
        keys = keys_dict[k1]
        array = output_dict[k1]
        for k2, v in zip(keys, array.T):
            d[k2] = v
        df = pd.DataFrame(d, index=stages)
        df.index.name = 'Position'
        if coordinate_frame is not None:
            df = pd.concat([coordinate_frame.copy(), df.reset_index(drop=True)], axis=1)
            df.index = stages
            df.index.name = 'Position'
        dfs_dict[k1] = df
    return dfs_dict


def build_profile_coordinate_frame(z, total_packed_height_m=None, beds=1):
    position = np.asarray(z, dtype=float)
    data = {"Position": position}
    if total_packed_height_m is not None:
        total_height = float(total_packed_height_m)
        beds = max(int(beds or 1), 1)
        single_bed_height = total_height / beds if beds else total_height
        height_m = position * total_height
        bed_id = np.floor(height_m / single_bed_height).astype(int) + 1
        bed_id = np.clip(bed_id, 1, beds)
        bed_position_m = height_m - (bed_id - 1) * single_bed_height
        bed_position_m = np.clip(bed_position_m, 0.0, single_bed_height)
        data.update(
            {
                "height_m": height_m,
                "bed_id": bed_id,
                "bed_position_m": bed_position_m,
            }
        )
    return pd.DataFrame(data)


def save_run_outputs(
    Y_scaled,
    z,
    parameters,
    save_run_results=True,
    plot_temperature=False,
    profile_metadata=None,
    include_coordinate_columns=False,
    legacy_excel=True,
):
    n = len(z)
    outputs_0, keys_dict = abs_column(z[0], Y_scaled.T[0], parameters, run_type='saving', column_names=True)
    sheetnames = list(keys_dict.keys())

    # Initialize output arrays
    output_dict = {k: np.zeros((n, len(outputs_0[k]))) for k in sheetnames}

    # Populate output arrays
    for i in range(n):
        outputs, _ = abs_column(z[i], Y_scaled.T[i], parameters, run_type='saving')
        for k in sheetnames:
            output_dict[k][i] = outputs[k]

    coordinate_frame = None
    if include_coordinate_columns:
        profile_metadata = profile_metadata or {}
        coordinate_frame = build_profile_coordinate_frame(
            z,
            total_packed_height_m=profile_metadata.get("total_packed_height_m"),
            beds=profile_metadata.get("beds", 1),
        )

    # Convert to DataFrame dict
    dfs_dict = make_dfs_dict(output_dict, keys_dict, z, coordinate_frame=coordinate_frame)

    if save_run_results and legacy_excel:
        # Locate or create the Excel workbook
        base = os.path.dirname(__file__)
        results_dir = os.path.abspath(os.path.join(base, '..', 'data', 'Results'))
        os.makedirs(results_dir, exist_ok=True)
        path = os.path.join(results_dir, 'Profiles.xlsx')

        if os.path.exists(path):
            wb = load_workbook(path)
        else:
            wb = Workbook()
            default = wb.active
            wb.remove(default)

        # Remove sheets not in current run
        existing = set(wb.sheetnames)
        wanted = set(dfs_dict.keys())
        for name in existing - wanted:
            del wb[name]

        # Create/populate sheets
        for sheetname, df in dfs_dict.items():
            if sheetname in wb.sheetnames:
                del wb[sheetname]
            ws = wb.create_sheet(title=sheetname)
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            # Freeze top row
            ws.freeze_panes = 'A2'

        # Reorder sheets
        wb._sheets = [wb[name] for name in dfs_dict.keys()]

        # Save workbook
        wb.save(path)

    return dfs_dict


def write_profile_csvs(dfs_dict, output_dir, metadata=None):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    metadata = dict(metadata or {})
    files = []
    for sheetname, df in dfs_dict.items():
        csv_df = _profile_csv_frame(df)
        path = output_dir / f"{sheetname}.csv"
        csv_df.to_csv(path, index=False)
        files.append(path.name)

    metadata["profile_csv_files"] = files
    metadata.setdefault("profile_csv_status", "written" if files else "empty")
    metadata.setdefault("profile_csv_dir", str(output_dir))
    (output_dir / "profile_manifest.json").write_text(
        json.dumps(_json_safe(metadata), indent=2, sort_keys=True),
        encoding="utf-8",
    )
    manifest_row = {
        key: ";".join(value) if isinstance(value, list) else value
        for key, value in metadata.items()
    }
    pd.DataFrame([manifest_row]).to_csv(output_dir / "profile_manifest.csv", index=False)
    return {
        "profile_csv_dir": str(output_dir),
        "profile_csv_status": metadata["profile_csv_status"],
        "profile_csv_files": ";".join(files),
    }


def _profile_csv_frame(df):
    if "Position" in df.columns:
        return df.copy()
    return df.reset_index()


def _json_safe(value):
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    if hasattr(value, "item"):
        return _json_safe(value.item())
    if isinstance(value, float) and not np.isfinite(value):
        return None
    return value
